# photovault/routes/admin_export.py
from flask import Blueprint, request, redirect, url_for, flash, make_response, current_app
from flask_login import login_required, current_user
from photovault import db
from photovault.models import User, Photo
from datetime import datetime
import csv
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import logging

logger = logging.getLogger(__name__)

admin_export_bp = Blueprint('admin_export', __name__)

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not (current_user.is_admin or current_user.is_superuser):
            flash('Access denied. Admin privileges required.', 'danger')
            return redirect(url_for('main.home'))
        return f(*args, **kwargs)
    return decorated_function

def superuser_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_superuser:
            flash('Access denied. Superuser privileges required.', 'danger')
            return redirect(url_for('main.home'))
        return f(*args, **kwargs)
    return decorated_function

@admin_export_bp.route('/export/users/csv')
@login_required
@admin_required
def export_users_csv():
    """Export all users to CSV file"""
    users = User.query.order_by(User.created_at.desc()).all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow([
        'ID', 'Username', 'Email', 'Is Admin', 'Is Superuser', 
        'Created At', 'Last Login', 'Total Photos', 'Edited Photos', 
        'Storage Used (MB)', 'Account Status'
    ])
    
    for user in users:
        user_photos = Photo.query.filter_by(user_id=user.id).all()
        total_photos = len(user_photos)
        edited_photos = sum(1 for photo in user_photos if photo.edited_filename)
        total_size = sum(photo.file_size or 0 for photo in user_photos)
        storage_mb = round(total_size / (1024 * 1024), 2) if total_size > 0 else 0
        
        writer.writerow([
            user.id,
            user.username,
            user.email,
            'Yes' if user.is_admin else 'No',
            'Yes' if user.is_superuser else 'No',
            user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else 'N/A',
            user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else 'Never',
            total_photos,
            edited_photos,
            storage_mb,
            'Active'
        ])
    
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename=photovault_users_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    response.headers['Content-Type'] = 'text/csv'
    
    logger.info(f"CSV export of {len(users)} users by admin {current_user.username}")
    return response

@admin_export_bp.route('/export/users/excel')
@login_required
@admin_required
def export_users_excel():
    """Export all users to Excel file"""
    users = User.query.order_by(User.created_at.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    headers = [
        'ID', 'Username', 'Email', 'Is Admin', 'Is Superuser', 
        'Created At', 'Last Login', 'Total Photos', 'Edited Photos', 
        'Storage Used (MB)', 'Account Status'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for row_num, user in enumerate(users, 2):
        user_photos = Photo.query.filter_by(user_id=user.id).all()
        total_photos = len(user_photos)
        edited_photos = sum(1 for photo in user_photos if photo.edited_filename)
        total_size = sum(photo.file_size or 0 for photo in user_photos)
        storage_mb = round(total_size / (1024 * 1024), 2) if total_size > 0 else 0
        
        ws.cell(row=row_num, column=1, value=user.id)
        ws.cell(row=row_num, column=2, value=user.username)
        ws.cell(row=row_num, column=3, value=user.email)
        ws.cell(row=row_num, column=4, value='Yes' if user.is_admin else 'No')
        ws.cell(row=row_num, column=5, value='Yes' if user.is_superuser else 'No')
        ws.cell(row=row_num, column=6, value=user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else 'N/A')
        ws.cell(row=row_num, column=7, value=user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else 'Never')
        ws.cell(row=row_num, column=8, value=total_photos)
        ws.cell(row=row_num, column=9, value=edited_photos)
        ws.cell(row=row_num, column=10, value=storage_mb)
        ws.cell(row=row_num, column=11, value='Active')
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename=photovault_users_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    logger.info(f"Excel export of {len(users)} users by admin {current_user.username}")
    return response

@admin_export_bp.route('/batch/delete', methods=['POST'])
@login_required
@superuser_required
def batch_delete_users():
    """Batch delete multiple users"""
    user_ids = request.form.getlist('user_ids[]')
    
    if not user_ids:
        flash('No users selected for deletion.', 'warning')
        return redirect(url_for('admin.dashboard'))
    
    deleted_count = 0
    errors = []
    
    for user_id in user_ids:
        try:
            user = User.query.get(int(user_id))
            if not user:
                continue
                
            if user.is_superuser:
                errors.append(f"Cannot delete superuser: {user.username}")
                continue
            
            if user.id == current_user.id:
                errors.append("Cannot delete your own account")
                continue
            
            for photo in user.photos:
                filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], photo.filename)
                if os.path.exists(filepath):
                    try:
                        os.remove(filepath)
                    except OSError:
                        pass
                
                if photo.edited_filename:
                    edited_filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], photo.edited_filename)
                    if os.path.exists(edited_filepath):
                        try:
                            os.remove(edited_filepath)
                        except OSError:
                            pass
            
            db.session.delete(user)
            deleted_count += 1
            
        except Exception as e:
            errors.append(f"Error deleting user ID {user_id}: {str(e)}")
    
    db.session.commit()
    
    if deleted_count > 0:
        flash(f'Successfully deleted {deleted_count} user(s).', 'success')
    
    for error in errors:
        flash(error, 'danger')
    
    logger.info(f"Batch deletion: {deleted_count} users deleted by {current_user.username}")
    return redirect(url_for('admin.dashboard'))

@admin_export_bp.route('/batch/toggle-admin', methods=['POST'])
@login_required
@superuser_required
def batch_toggle_admin():
    """Batch toggle admin status for multiple users"""
    user_ids = request.form.getlist('user_ids[]')
    action = request.form.get('action', 'grant')
    
    if not user_ids:
        flash('No users selected.', 'warning')
        return redirect(url_for('admin.dashboard'))
    
    updated_count = 0
    errors = []
    
    for user_id in user_ids:
        try:
            user = User.query.get(int(user_id))
            if not user:
                continue
            
            if user.is_superuser:
                errors.append(f"Cannot modify superuser: {user.username}")
                continue
            
            if action == 'grant':
                user.is_admin = True
            else:
                user.is_admin = False
            
            updated_count += 1
            
        except Exception as e:
            errors.append(f"Error updating user ID {user_id}: {str(e)}")
    
    db.session.commit()
    
    if updated_count > 0:
        action_text = 'granted' if action == 'grant' else 'revoked'
        flash(f'Admin status {action_text} for {updated_count} user(s).', 'success')
    
    for error in errors:
        flash(error, 'danger')
    
    logger.info(f"Batch admin toggle: {updated_count} users updated by {current_user.username}")
    return redirect(url_for('admin.dashboard'))
