# photovault/routes/admin.py
from flask import Blueprint, render_template, redirect, url_for, flash, request, abort, current_app, jsonify, make_response
from flask_login import login_required, current_user
from sqlalchemy import func, case, text
from photovault import db
from photovault.models import User, Photo
from datetime import datetime, timedelta
import os
import logging
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Set up logging
logger = logging.getLogger(__name__)

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

def superuser_required(f):
    """Decorator to require superuser access"""
    def wrap(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_superuser:
            logger.warning(f"Unauthorized superuser access attempt by user {current_user.id if current_user.is_authenticated else 'anonymous'}")
            flash("You do not have permission to access that page.", "danger")
            abort(403)  # Forbidden
        return f(*args, **kwargs)
    wrap.__name__ = f.__name__  # Important for endpoint naming
    return wrap

def admin_required(f):
    """Decorator to require either admin or superuser access"""
    def wrap(*args, **kwargs):
        if not current_user.is_authenticated or not (current_user.is_admin or current_user.is_superuser):
            logger.warning(f"Unauthorized admin access attempt by user {current_user.id if current_user.is_authenticated else 'anonymous'}")
            flash("You do not have permission to access that page.", "danger")
            abort(403)  # Forbidden
        return f(*args, **kwargs)
    wrap.__name__ = f.__name__
    return wrap

@admin_bp.route('/')
@admin_bp.route('/dashboard')
@login_required
@admin_required
def dashboard():
    """Enhanced admin dashboard showing user list with statistics"""
    # Get all users
    users = User.query.order_by(User.created_at.desc()).all()
    
    # Calculate statistics for each user manually (no complex SQLAlchemy)
    users_with_stats = []
    total_storage_all = 0
    total_photos_all = 0
    total_edited_all = 0
    
    for user in users:
        # Get user's photos
        user_photos = Photo.query.filter_by(user_id=user.id).all()
        
        # Calculate stats manually
        total_photos = len(user_photos)
        edited_photos = 0
        total_size = 0
        
        for photo in user_photos:
            if photo.edited_filename is not None:
                edited_photos += 1
            if photo.file_size:
                total_size += photo.file_size
        
        # Add to totals
        total_storage_all += total_size
        total_photos_all += total_photos
        total_edited_all += edited_photos
        
        # Create tuple matching the expected format
        users_with_stats.append((user, total_photos, edited_photos, total_size))
    
    # Overall statistics
    total_users = len(users)
    
    stats = {
        'total_users': total_users,
        'total_photos': total_photos_all,
        'total_edited': total_edited_all,
        'total_storage': total_storage_all,
        'total_storage_mb': round(total_storage_all / (1024 * 1024), 2) if total_storage_all else 0
    }
    
    return render_template('admin/dashboard.html', users_with_stats=users_with_stats, stats=stats)

@admin_bp.route('/api/statistics')
@login_required
@admin_required
def api_statistics():
    """JSON API endpoint for dashboard statistics"""
    # Get all users
    users = User.query.order_by(User.created_at.desc()).all()
    
    # Calculate statistics for each user manually 
    total_storage_all = 0
    total_photos_all = 0
    total_edited_all = 0
    
    for user in users:
        # Get user's photos
        user_photos = Photo.query.filter_by(user_id=user.id).all()
        
        # Calculate stats manually
        total_photos = len(user_photos)
        edited_photos = 0
        total_size = 0
        
        for photo in user_photos:
            if photo.edited_filename is not None:
                edited_photos += 1
            if photo.file_size:
                total_size += photo.file_size
        
        # Add to totals
        total_storage_all += total_size
        total_photos_all += total_photos
        total_edited_all += edited_photos
    
    # Overall statistics
    total_users = len(users)
    
    stats = {
        'total_users': total_users,
        'total_photos': total_photos_all,
        'total_edited': total_edited_all,
        'total_storage': total_storage_all,
        'total_storage_mb': round(total_storage_all / (1024 * 1024), 2) if total_storage_all else 0
    }
    
    return jsonify(stats)

@admin_bp.route('/user/<int:user_id>')
@login_required
@admin_required
def user_detail(user_id):
    """View detailed information about a specific user"""
    user = User.query.get_or_404(user_id)
    
    # Get user's photos
    photos = Photo.query.filter_by(user_id=user_id).order_by(Photo.created_at.desc()).all()
    
    # Calculate statistics manually
    total_photos = len(photos)
    edited_photos = 0
    total_size = 0
    
    for photo in photos:
        if photo.edited_filename is not None:
            edited_photos += 1
        if photo.file_size:
            total_size += photo.file_size
    
    user_stats = {
        'total_photos': total_photos,
        'edited_photos': edited_photos,
        'original_photos': total_photos - edited_photos,
        'total_size': total_size,
        'total_size_mb': round(total_size / (1024 * 1024), 2) if total_size > 0 else 0,
        'avg_file_size': round(total_size / total_photos / 1024, 2) if total_photos > 0 else 0  # in KB
    }
    
    return render_template('admin/user_detail.html', user=user, photos=photos, user_stats=user_stats)

@admin_bp.route('/user/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id):
    """Edit user information"""
    user = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        new_username = request.form.get('username', '').strip()
        new_email = request.form.get('email', '').strip()
        
        # Validate input
        if not new_username or not new_email:
            flash("Username and email are required.", "danger")
            return redirect(url_for('admin.edit_user', user_id=user_id))
        
        # Check if username or email is taken by another user
        existing_user = User.query.filter(
            User.id != user_id,
            (User.username == new_username) | (User.email == new_email)
        ).first()
        
        if existing_user:
            flash("Username or email already exists.", "danger")
            return redirect(url_for('admin.edit_user', user_id=user_id))
        
        # Update user information
        user.username = new_username
        user.email = new_email
        
        db.session.commit()
        flash(f"User information updated successfully.", "success")
        return redirect(url_for('admin.user_detail', user_id=user_id))
    
    return render_template('admin/edit_user.html', user=user)

@admin_bp.route('/user/<int:user_id>/reset-password', methods=['POST'])
@login_required
@superuser_required  # Only superusers can reset passwords
def reset_user_password(user_id):
    """Reset a user's password"""
    user = User.query.get_or_404(user_id)
    
    new_password = request.form.get('new_password', '').strip()
    if not new_password:
        flash("New password is required.", "danger")
        return redirect(url_for('admin.user_detail', user_id=user_id))
    
    if len(new_password) < 6:
        flash("Password must be at least 6 characters long.", "danger")
        return redirect(url_for('admin.user_detail', user_id=user_id))
    
    user.set_password(new_password)
    db.session.commit()
    
    flash(f"Password reset for user {user.username}.", "success")
    return redirect(url_for('admin.user_detail', user_id=user_id))

@admin_bp.route('/users/toggle_admin/<int:user_id>', methods=['POST'])
@login_required
@superuser_required # Only superusers can toggle admin status
def toggle_admin(user_id):
    """Toggle the admin status of a user"""
    user = User.query.get_or_404(user_id)
    
    # Prevent users from modifying superusers unless they are also superusers
    if user.is_superuser and not current_user.is_superuser:
        flash("You cannot modify superuser accounts.", "danger")
        return redirect(url_for('admin.dashboard'))
        
    user.is_admin = not user.is_admin
    db.session.commit()
    status = "granted" if user.is_admin else "revoked"
    flash(f"Admin status {status} for user {user.username}.", "success")
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/users/toggle_superuser/<int:user_id>', methods=['POST'])
@login_required
@superuser_required # Only superusers can toggle superuser status
def toggle_superuser(user_id):
    """Toggle the superuser status of a user"""
    user = User.query.get_or_404(user_id)
    
    # Critical: Prevent a superuser from removing their own superuser status
    if user.id == current_user.id:
        flash("You cannot change your own superuser status.", "warning")
        return redirect(url_for('admin.dashboard'))

    user.is_superuser = not user.is_superuser
    db.session.commit()
    status = "granted" if user.is_superuser else "revoked"
    flash(f"Superuser status {status} for user {user.username}.", "success")
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/users/delete/<int:user_id>', methods=['POST'])
@login_required
@superuser_required # Only superusers can delete users
def delete_user(user_id):
    """Delete a user and their photos"""
    user = User.query.get_or_404(user_id)
    
    # Prevent users from deleting superusers
    if user.is_superuser:
        flash("Superuser accounts cannot be deleted.", "danger")
        return redirect(url_for('admin.dashboard'))
        
    # Prevent users from deleting themselves
    if user.id == current_user.id:
        flash("You cannot delete your own account.", "danger")
        return redirect(url_for('admin.dashboard'))

    username = user.username
    
    # Delete user's photos from the filesystem
    for photo in user.photos:
        # Delete original file
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], photo.filename)
        if os.path.exists(filepath):
            try:
                os.remove(filepath)
            except OSError as e:
                print(f"Error deleting file {filepath}: {e}")
        
        # Delete edited file if it exists
        if photo.edited_filename:
            edited_filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], photo.edited_filename)
            if os.path.exists(edited_filepath):
                try:
                    os.remove(edited_filepath)
                except OSError as e:
                    print(f"Error deleting edited file {edited_filepath}: {e}")
    
    db.session.delete(user)
    db.session.commit()
    flash(f"User {username} and all their photos deleted successfully.", "success")
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/photo/<int:photo_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_photo(photo_id):
    """Delete a specific photo (admin can delete any photo)"""
    photo = Photo.query.get_or_404(photo_id)
    user_id = photo.user_id
    
    # Delete files from filesystem
    original_filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], photo.filename)
    if os.path.exists(original_filepath):
        try:
            os.remove(original_filepath)
        except OSError as e:
            print(f"Error deleting original file {original_filepath}: {e}")
    
    if photo.edited_filename:
        edited_filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], photo.edited_filename)
        if os.path.exists(edited_filepath):
            try:
                os.remove(edited_filepath)
            except OSError as e:
                print(f"Error deleting edited file {edited_filepath}: {e}")
    
    db.session.delete(photo)
    db.session.commit()
    
    flash("Photo deleted successfully.", "success")
    return redirect(url_for('admin.user_detail', user_id=user_id))

@admin_bp.route('/statistics')
@login_required
@admin_required
def statistics():
    """View detailed system statistics with simple, reliable queries"""
    try:
        # Simple user statistics - avoid complex aggregations
        all_users = User.query.all()
        total_users = len(all_users)
        admin_users = sum(1 for user in all_users if user.is_admin)
        superusers = sum(1 for user in all_users if user.is_superuser)
        
        # Recent users (last 30 days)
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        recent_users = 0
        for user in all_users:
            if user.created_at and user.created_at >= thirty_days_ago:
                recent_users += 1
        
        # Simple photo statistics
        all_photos = Photo.query.all()
        total_photos = len(all_photos)
        edited_photos = 0
        total_size = 0
        recent_uploads = 0
        
        for photo in all_photos:
            if photo.edited_filename is not None:
                edited_photos += 1
            if photo.file_size:
                total_size += photo.file_size
            if photo.created_at and photo.created_at >= thirty_days_ago:
                recent_uploads += 1
        
        # Most active users - simple approach
        user_photo_counts = {}
        for photo in all_photos:
            user_id = photo.user_id
            if user_id not in user_photo_counts:
                user_photo_counts[user_id] = 0
            user_photo_counts[user_id] += 1
        
        # Get top 10 most active users
        most_active_users = []
        for user in all_users:
            photo_count = user_photo_counts.get(user.id, 0)
            if photo_count > 0:
                most_active_users.append((user.username, photo_count))
        
        # Sort by photo count and take top 10
        most_active_users.sort(key=lambda x: x[1], reverse=True)
        most_active_users = most_active_users[:10]
        
        # Create statistics dictionary
        statistics = {
            'users': {
                'total': total_users,
                'admins': admin_users,
                'superusers': superusers,
                'recent': recent_users
            },
            'photos': {
                'total': total_photos,
                'edited': edited_photos,
                'original_only': total_photos - edited_photos,
                'recent_uploads': recent_uploads
            },
            'storage': {
                'total_bytes': total_size,
                'total_mb': round(total_size / (1024 * 1024), 2) if total_size > 0 else 0,
                'total_gb': round(total_size / (1024 * 1024 * 1024), 2) if total_size > 0 else 0,
                'avg_file_size_kb': round(total_size / total_photos / 1024, 2) if total_photos > 0 else 0
            },
            'most_active_users': most_active_users
        }
        
        logger.info(f"Statistics loaded successfully by admin {current_user.username}")
        print(f"Statistics data: {statistics}")  # Debug output
        
        return render_template('admin/statistics.html', stats=statistics)
    
    except Exception as e:
        logger.error(f"Error loading statistics: {str(e)}")
        print(f"Statistics error: {str(e)}")  # Debug output
        
        # Provide fallback empty statistics
        fallback_stats = {
            'users': {'total': 0, 'admins': 0, 'superusers': 0, 'recent': 0},
            'photos': {'total': 0, 'edited': 0, 'original_only': 0, 'recent_uploads': 0},
            'storage': {'total_bytes': 0, 'total_mb': 0, 'total_gb': 0, 'avg_file_size_kb': 0},
            'most_active_users': []
        }
        
        flash("Error loading statistics. Showing basic information.", "warning")
        return render_template('admin/statistics.html', stats=fallback_stats)

@admin_bp.route('/profile')
@login_required
@admin_required # Both admins and superusers can view their profile
def profile():
    """View admin/superuser profile"""
    return render_template('admin/profile.html', user=current_user)

@admin_bp.route('/export/users/csv')
@login_required
@admin_required
def export_users_csv():
    """Export all users to CSV file"""
    # Get all users with their statistics
    users = User.query.order_by(User.created_at.desc()).all()
    
    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'ID', 'Username', 'Email', 'Is Admin', 'Is Superuser', 
        'Created At', 'Last Login', 'Total Photos', 'Edited Photos', 
        'Storage Used (MB)', 'Account Status'
    ])
    
    # Write user data
    for user in users:
        # Calculate user statistics
        user_photos = Photo.query.filter_by(user_id=user.id).all()
        total_photos = len(user_photos)
        edited_photos = sum(1 for photo in user_photos if photo.edited_filename)
        total_size = sum(photo.file_size or 0 for photo in user_photos)
        storage_mb = round(total_size / (1024 * 1024), 2) if total_size > 0 else 0
        
        writer.writerow([
            user.id,
            user.username,
            user.email,
            'Yes' if user.is_admin else 'No',
            'Yes' if user.is_superuser else 'No',
            user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else 'N/A',
            user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else 'Never',
            total_photos,
            edited_photos,
            storage_mb,
            'Active'
        ])
    
    # Create response
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename=photovault_users_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    response.headers['Content-Type'] = 'text/csv'
    
    logger.info(f"CSV export of {len(users)} users by admin {current_user.username}")
    return response

@admin_bp.route('/export/users/excel')
@login_required
@admin_required
def export_users_excel():
    """Export all users to Excel file"""
    # Get all users with their statistics
    users = User.query.order_by(User.created_at.desc()).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    
    # Define header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write header
    headers = [
        'ID', 'Username', 'Email', 'Is Admin', 'Is Superuser', 
        'Created At', 'Last Login', 'Total Photos', 'Edited Photos', 
        'Storage Used (MB)', 'Account Status'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write user data
    for row_num, user in enumerate(users, 2):
        # Calculate user statistics
        user_photos = Photo.query.filter_by(user_id=user.id).all()
        total_photos = len(user_photos)
        edited_photos = sum(1 for photo in user_photos if photo.edited_filename)
        total_size = sum(photo.file_size or 0 for photo in user_photos)
        storage_mb = round(total_size / (1024 * 1024), 2) if total_size > 0 else 0
        
        ws.cell(row=row_num, column=1, value=user.id)
        ws.cell(row=row_num, column=2, value=user.username)
        ws.cell(row=row_num, column=3, value=user.email)
        ws.cell(row=row_num, column=4, value='Yes' if user.is_admin else 'No')
        ws.cell(row=row_num, column=5, value='Yes' if user.is_superuser else 'No')
        ws.cell(row=row_num, column=6, value=user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else 'N/A')
        ws.cell(row=row_num, column=7, value=user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else 'Never')
        ws.cell(row=row_num, column=8, value=total_photos)
        ws.cell(row=row_num, column=9, value=edited_photos)
        ws.cell(row=row_num, column=10, value=storage_mb)
        ws.cell(row=row_num, column=11, value='Active')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename=photovault_users_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    logger.info(f"Excel export of {len(users)} users by admin {current_user.username}")
    return response

@admin_bp.route('/batch/delete', methods=['POST'])
@login_required
@superuser_required
def batch_delete_users():
    """Batch delete multiple users"""
    user_ids = request.form.getlist('user_ids[]')
    
    if not user_ids:
        flash('No users selected for deletion.', 'warning')
        return redirect(url_for('admin.dashboard'))
    
    deleted_count = 0
    errors = []
    
    for user_id in user_ids:
        try:
            user = User.query.get(int(user_id))
            if not user:
                continue
                
            # Prevent deletion of superusers and self
            if user.is_superuser:
                errors.append(f"Cannot delete superuser: {user.username}")
                continue
            
            if user.id == current_user.id:
                errors.append("Cannot delete your own account")
                continue
            
            # Delete user's photos from filesystem
            for photo in user.photos:
                filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], photo.filename)
                if os.path.exists(filepath):
                    try:
                        os.remove(filepath)
                    except OSError:
                        pass
                
                if photo.edited_filename:
                    edited_filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], photo.edited_filename)
                    if os.path.exists(edited_filepath):
                        try:
                            os.remove(edited_filepath)
                        except OSError:
                            pass
            
            # Delete user
            db.session.delete(user)
            deleted_count += 1
            
        except Exception as e:
            errors.append(f"Error deleting user ID {user_id}: {str(e)}")
    
    # Commit all deletions
    db.session.commit()
    
    # Show results
    if deleted_count > 0:
        flash(f'Successfully deleted {deleted_count} user(s).', 'success')
    
    for error in errors:
        flash(error, 'danger')
    
    logger.info(f"Batch deletion: {deleted_count} users deleted by {current_user.username}")
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/batch/toggle-admin', methods=['POST'])
@login_required
@superuser_required
def batch_toggle_admin():
    """Batch toggle admin status for multiple users"""
    user_ids = request.form.getlist('user_ids[]')
    action = request.form.get('action', 'grant')  # 'grant' or 'revoke'
    
    if not user_ids:
        flash('No users selected.', 'warning')
        return redirect(url_for('admin.dashboard'))
    
    updated_count = 0
    errors = []
    
    for user_id in user_ids:
        try:
            user = User.query.get(int(user_id))
            if not user:
                continue
            
            # Prevent modifying superusers
            if user.is_superuser:
                errors.append(f"Cannot modify superuser: {user.username}")
                continue
            
            # Update admin status
            if action == 'grant':
                user.is_admin = True
            else:
                user.is_admin = False
            
            updated_count += 1
            
        except Exception as e:
            errors.append(f"Error updating user ID {user_id}: {str(e)}")
    
    # Commit all changes
    db.session.commit()
    
    # Show results
    if updated_count > 0:
        action_text = 'granted' if action == 'grant' else 'revoked'
        flash(f'Admin status {action_text} for {updated_count} user(s).', 'success')
    
    for error in errors:
        flash(error, 'danger')
    
    logger.info(f"Batch admin toggle: {updated_count} users updated by {current_user.username}")
    return redirect(url_for('admin.dashboard'))